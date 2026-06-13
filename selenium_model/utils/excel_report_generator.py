"""
Master Test Report Generator
Generates comprehensive Excel report for all testing and audit results
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import json

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate comprehensive Excel test report"""
    
    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.wb = Workbook()
        self.ws_list = {}
        
        # Style definitions
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.passed_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.failed_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.failed_font = Font(color="FFFFFF", bold=True)
        self.medium_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.low_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _style_header(self, sheet, row_num=1):
        """Style header row"""
        for col in sheet.iter_cols(min_row=row_num, max_row=row_num):
            for cell in col:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border
    
    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_executive_summary(self, summary_data: Dict):
        """Create Executive Summary sheet"""
        ws = self.wb.active
        ws.title = "Executive Summary"
        self.ws_list["Executive Summary"] = ws
        
        # Title
        ws['A1'] = "EVENTBRIDGE PROJECT - TEST AUDIT REPORT"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Scan Date
        ws['A3'] = "Scan Date:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Project Info
        row = 5
        for key, value in summary_data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def create_functional_test_results(self, test_results: List[Dict]):
        """Create Functional Test Results sheet"""
        ws = self.wb.create_sheet("Functional Test Results")
        self.ws_list["Functional Test Results"] = ws
        
        # Headers
        headers = ["Test ID", "Module", "Scenario", "Expected Result", "Actual Result", 
                  "Status", "Execution Time (s)", "Screenshot Path"]
        ws.append(headers)
        self._style_header(ws)
        
        # Data
        for result in test_results:
            status_color = self.passed_fill if result.get("status") == "PASSED" else self.failed_fill
            
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = result.get("test_id", "")
            ws[f'B{row_num}'] = result.get("module", "")
            ws[f'C{row_num}'] = result.get("scenario", "")
            ws[f'D{row_num}'] = result.get("expected_result", "")
            ws[f'E{row_num}'] = result.get("actual_result", "")
            ws[f'F{row_num}'] = result.get("status", "")
            ws[f'G{row_num}'] = result.get("execution_time", 0)
            ws[f'H{row_num}'] = result.get("screenshot_path", "")
            
            # Color status cell
            ws[f'F{row_num}'].fill = status_color
            if result.get("status") == "FAILED":
                ws[f'F{row_num}'].font = self.failed_font
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_functional_coverage(self, coverage_data: List[Dict]):
        """Create Functional Coverage sheet"""
        ws = self.wb.create_sheet("Functional Coverage")
        self.ws_list["Functional Coverage"] = ws
        
        headers = ["Page", "Functionality", "Coverage Status", "Remarks"]
        ws.append(headers)
        self._style_header(ws)
        
        for item in coverage_data:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = item.get("page", "")
            ws[f'B{row_num}'] = item.get("functionality", "")
            ws[f'C{row_num}'] = item.get("coverage_status", "")
            ws[f'D{row_num}'] = item.get("remarks", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_defect_report(self, defects: List[Dict]):
        """Create Defect Report sheet"""
        ws = self.wb.create_sheet("Defect Report")
        self.ws_list["Defect Report"] = ws
        
        headers = ["Bug ID", "Module", "Description", "Steps to Reproduce", 
                  "Severity", "Evidence", "Status"]
        ws.append(headers)
        self._style_header(ws)
        
        for defect in defects:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = defect.get("bug_id", "")
            ws[f'B{row_num}'] = defect.get("module", "")
            ws[f'C{row_num}'] = defect.get("description", "")
            ws[f'D{row_num}'] = defect.get("steps_to_reproduce", "")
            ws[f'E{row_num}'] = defect.get("severity", "")
            ws[f'F{row_num}'] = defect.get("evidence", "")
            ws[f'G{row_num}'] = defect.get("status", "")
            
            # Color severity
            if defect.get("severity") == "High":
                ws[f'E{row_num}'].fill = self.high_fill
            elif defect.get("severity") == "Medium":
                ws[f'E{row_num}'].fill = self.medium_fill
            elif defect.get("severity") == "Low":
                ws[f'E{row_num}'].fill = self.low_fill
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_unused_files(self, unused_files: List[Dict]):
        """Create Unused Files sheet"""
        ws = self.wb.create_sheet("Unused Files")
        self.ws_list["Unused Files"] = ws
        
        headers = ["File Name", "Path", "Reason", "Severity"]
        ws.append(headers)
        self._style_header(ws)
        
        for file_item in unused_files:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = file_item.get("file_name", "")
            ws[f'B{row_num}'] = file_item.get("path", "")
            ws[f'C{row_num}'] = file_item.get("reason", "")
            ws[f'D{row_num}'] = file_item.get("severity", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_dead_code(self, dead_code_items: List[Dict]):
        """Create Dead Code sheet"""
        ws = self.wb.create_sheet("Dead Code")
        self.ws_list["Dead Code"] = ws
        
        headers = ["File", "Function or Class", "Line Number", "Recommendation"]
        ws.append(headers)
        self._style_header(ws)
        
        for item in dead_code_items:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = item.get("file", "")
            ws[f'B{row_num}'] = item.get("function_or_class", "")
            ws[f'C{row_num}'] = item.get("line_number", "")
            ws[f'D{row_num}'] = item.get("recommendation", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_broken_links(self, broken_links: List[Dict]):
        """Create Broken Links sheet"""
        ws = self.wb.create_sheet("Broken Links")
        self.ws_list["Broken Links"] = ws
        
        headers = ["URL", "Source Page", "Status Code", "Result"]
        ws.append(headers)
        self._style_header(ws)
        
        for link in broken_links:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = link.get("url", "")
            ws[f'B{row_num}'] = link.get("source_page", "")
            ws[f'C{row_num}'] = link.get("status_code", "")
            ws[f'D{row_num}'] = link.get("result", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_accessibility_findings(self, accessibility_items: List[Dict]):
        """Create Accessibility Findings sheet"""
        ws = self.wb.create_sheet("Accessibility Findings")
        self.ws_list["Accessibility Findings"] = ws
        
        headers = ["Page", "Issue", "Severity", "Recommendation"]
        ws.append(headers)
        self._style_header(ws)
        
        for item in accessibility_items:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = item.get("page", "")
            ws[f'B{row_num}'] = item.get("issue", "")
            ws[f'C{row_num}'] = item.get("severity", "")
            ws[f'D{row_num}'] = item.get("recommendation", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_api_validation(self, api_results: List[Dict]):
        """Create API Validation Results sheet"""
        ws = self.wb.create_sheet("API Validation Results")
        self.ws_list["API Validation Results"] = ws
        
        headers = ["Endpoint", "Method", "Expected Status", "Actual Status", "Result"]
        ws.append(headers)
        self._style_header(ws)
        
        for result in api_results:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = result.get("endpoint", "")
            ws[f'B{row_num}'] = result.get("method", "")
            ws[f'C{row_num}'] = result.get("expected_status", "")
            ws[f'D{row_num}'] = result.get("actual_status", "")
            ws[f'E{row_num}'] = result.get("result", "")
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_ui_validation(self, ui_findings: List[Dict]):
        """Create UI Validation Findings sheet"""
        ws = self.wb.create_sheet("UI Validation Findings")
        self.ws_list["UI Validation Findings"] = ws
        
        headers = ["Page", "Issue", "Severity", "Evidence"]
        ws.append(headers)
        self._style_header(ws)
        
        for item in ui_findings:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = item.get("page", "")
            ws[f'B{row_num}'] = item.get("issue", "")
            ws[f'C{row_num}'] = item.get("severity", "")
            ws[f'D{row_num}'] = item.get("evidence", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_performance_observations(self, performance_data: List[Dict]):
        """Create Performance Observations sheet"""
        ws = self.wb.create_sheet("Performance Observations")
        self.ws_list["Performance Observations"] = ws
        
        headers = ["Page", "Load Time (s)", "Observation", "Recommendation"]
        ws.append(headers)
        self._style_header(ws)
        
        for item in performance_data:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = item.get("page", "")
            ws[f'B{row_num}'] = item.get("load_time", "")
            ws[f'C{row_num}'] = item.get("observation", "")
            ws[f'D{row_num}'] = item.get("recommendation", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_user_journeys(self, user_journeys: List[Dict]):
        """Create User Journey Results sheet"""
        ws = self.wb.create_sheet("User Journey Results")
        self.ws_list["User Journey Results"] = ws
        
        headers = ["Journey Name", "Steps", "Result", "Evidence"]
        ws.append(headers)
        self._style_header(ws)
        
        for journey in user_journeys:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = journey.get("journey_name", "")
            ws[f'B{row_num}'] = journey.get("steps", "")
            ws[f'C{row_num}'] = journey.get("result", "")
            ws[f'D{row_num}'] = journey.get("evidence", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_security_observations(self, security_items: List[Dict]):
        """Create Security Observations sheet"""
        ws = self.wb.create_sheet("Security Observations")
        self.ws_list["Security Observations"] = ws
        
        headers = ["Area", "Observation", "Severity", "Recommendation"]
        ws.append(headers)
        self._style_header(ws)
        
        for item in security_items:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = item.get("area", "")
            ws[f'B{row_num}'] = item.get("observation", "")
            ws[f'C{row_num}'] = item.get("severity", "")
            ws[f'D{row_num}'] = item.get("recommendation", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_code_health(self, code_health_items: List[Dict]):
        """Create Code Health Summary sheet"""
        ws = self.wb.create_sheet("Code Health Summary")
        self.ws_list["Code Health Summary"] = ws
        
        headers = ["Category", "Finding", "Severity", "Recommendation"]
        ws.append(headers)
        self._style_header(ws)
        
        for item in code_health_items:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = item.get("category", "")
            ws[f'B{row_num}'] = item.get("finding", "")
            ws[f'C{row_num}'] = item.get("severity", "")
            ws[f'D{row_num}'] = item.get("recommendation", "")
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def create_recommendations(self, recommendations: List[Dict]):
        """Create Recommendations sheet"""
        ws = self.wb.create_sheet("Recommendations")
        self.ws_list["Recommendations"] = ws
        
        headers = ["Priority", "Recommendation", "Business Impact"]
        ws.append(headers)
        self._style_header(ws)
        
        for rec in recommendations:
            row_num = ws.max_row + 1
            ws[f'A{row_num}'] = rec.get("priority", "")
            ws[f'B{row_num}'] = rec.get("recommendation", "")
            ws[f'C{row_num}'] = rec.get("business_impact", "")
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row_num}'].border = self.border
        
        self._auto_adjust_columns(ws)
    
    def save(self):
        """Save workbook"""
        self.wb.save(self.output_path)
        logger.info(f"Excel report generated: {self.output_path}")
        return self.output_path
